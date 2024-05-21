import os
import time
from datetime import datetime
import openpyxl as xl
import pygame as pg

pg.init()
pg.mixer.init()

current_datetime = datetime.now()
current_month = current_datetime.month

width, height = 1280, 800

screen = pg.display.set_mode((width, height), pg.FULLSCREEN)
# screen = pg.display.set_mode((width, height))

button_width = 200
button_height = 100

exit_box_rect = pg.Rect(1220, 10, 50, 50)

pg.display.set_caption("Digital Timecard")

YELLOW = (255, 255, 0)
BG_GREY = (192, 192, 192)
LIGHT_GREY = (224, 224, 224)
BLACK = (0, 0, 0)
WHITE = (255, 255, 255)
BROWN = (51, 25, 0)

font = pg.font.SysFont("Meiryo UI", 36)


def check_and_create_excel(file_name, staff_list):
    if not os.path.exists(file_name):
        wb = xl.Workbook()
        wb.save(file_name)

    db = xl.load_workbook(file_name)
    for staff in staff_list:
        if staff.sheet_name not in db.sheetnames:
            sheet = db.create_sheet(staff.sheet_name)
            sheet['A1'] = "Date"
            sheet['B1'] = "Clock In"
            sheet['C1'] = "Clock Out"
    db.save(file_name)


class Member:
    def __init__(self, name, rect_loc, button_loc, sheet_name):
        self.name = name
        self.button_rect = pg.Rect(rect_loc)
        self.button_loc = button_loc
        self.button_col = LIGHT_GREY if ':out' in self.name else YELLOW
        self.pressed_col = YELLOW
        self.sheet_name = sheet_name

    @classmethod
    def clock_in_out_message(cls, name, clean_time):
        draw_panel(screen)
        if name.split(':')[1] == 'in':
            if clean_time[:2] <= '10':
                text = font.render(f"{name.replace(':in', '')}、おはようございます。{clean_time}に出勤しました。", True, WHITE)
            elif clean_time[:2] <= '18':
                text = font.render(f"{name.replace(':in', '')}、こんにちは。{clean_time}に出勤しました。", True, WHITE)
            else:
                text = font.render(f"{name.replace(':in', '')}、こんばんは。{clean_time}に出勤しました。", True, WHITE)
        else:
            text = font.render(f"{name.replace(':out', '')}、お疲れ様でした。{clean_time}に退勤しました。", True, WHITE)
        save_staff_list()
        text_rect = text.get_rect(center=(width // 2, 50))
        screen.blit(text, text_rect)
        pg.display.flip()
        time.sleep(2)

    @classmethod
    def input_time(cls, name, sheet_name, clean_time, current_day, current_date, current_month):
        db = xl.load_workbook(f'tc2assets/{current_month}.xlsx')
        sheet = db[sheet_name]
        time_cell = sheet.cell(current_day + 1, 2 if ':in' in name else 3)
        time_cell.value = clean_time[:5]
        date_cell = sheet.cell(current_day + 1, 1)
        date_cell.value = current_date
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        db.save(f'tc2assets/{current_month}.xlsx')

    def button_click(self, mouse_pos, clean_time, current_day, current_date, current_month):
        if self.button_rect.collidepoint(mouse_pos):
            if ':out' in self.name:
                self.name = self.name.replace(':out', ':in')
                self.button_col = self.pressed_col
            else:
                self.name = self.name.replace(':in', ':out')
                self.button_col = LIGHT_GREY
            Member.input_time(self.name, self.sheet_name, clean_time, current_day, current_date, current_month)
            Member.clock_in_out_message(self.name, clean_time)


def draw_panel(surface):
    pg.draw.rect(surface, BG_GREY, (0, 0, 1280, 800), 0)
    pg.draw.rect(surface, BROWN, (0, 0, 1280, 100), 0)
    pg.draw.rect(surface, BLACK, exit_box_rect)
    text = font.render("X", True, WHITE)
    text_rect = text.get_rect(center=exit_box_rect.center)
    surface.blit(text, text_rect)


def update_display(current_date, clean_time):
    draw_panel(screen)
    text = font.render(f"{current_date} {clean_time}", True, WHITE)
    text_rect = text.get_rect(center=(width // 2, 50))
    screen.blit(text, text_rect)


def update_buttons():
    for staff in staff_list:
        pg.draw.rect(screen, staff.button_col, staff.button_rect)
        pg.draw.rect(screen, BLACK, staff.button_rect, 2)
        text = font.render(f"{staff.name.replace(':in', '').replace(':out', '')}", True, BLACK)
        text_rect = text.get_rect(center=(staff.button_loc[0] + button_width / 2, staff.button_loc[1] + button_height / 2))
        screen.blit(text, text_rect)
    pg.display.flip()


def save_staff_list():
    with open('tc2assets/staff_list.txt', 'w', encoding='utf-8') as file:
        for staff in staff_list:
            file.write(f"{staff.name}\n")


run = True

buttons = [
    (100, 130, 200, 100), (320, 130, 200, 100), (540, 130, 200, 100), (760, 130, 200, 100), (980, 130, 200, 100),
    (100, 250, 200, 100), (320, 250, 200, 100), (540, 250, 200, 100), (760, 250, 200, 100), (980, 250, 200, 100),
    (100, 370, 200, 100), (320, 370, 200, 100), (540, 370, 200, 100), (760, 370, 200, 100), (980, 370, 200, 100),
    (100, 490, 200, 100), (320, 490, 200, 100), (540, 490, 200, 100), (760, 490, 200, 100), (980, 490, 200, 100)
]

staff_list = []

try:
    with open('tc2assets/staff_list.txt', encoding='utf-8') as file:
        lines = file.readlines()
        for i, line in enumerate(lines):
            name = line.strip()
            button_rect = buttons[i]
            staff_member = Member(name=name, rect_loc=button_rect, button_loc=button_rect, sheet_name=str(name.replace(':in', '').replace(':out', '')))
            staff_list.append(staff_member)
except FileNotFoundError:
    print("Staff list file not found. Please ensure 'tc2assets/staff_list.txt' exists.")
    run = False

check_and_create_excel(f'tc2assets/{current_month}.xlsx', staff_list)

while run:
    current_datetime = datetime.now()
    time_frag = (str(current_datetime).split()[1])
    clean_time = time_frag[:8]
    current_date = current_datetime.date()
    current_month = current_datetime.month
    current_day = current_datetime.day
    update_display(current_date, clean_time)
    update_buttons()
    for event in pg.event.get():
        if event.type == pg.QUIT:
            run = False
        if event.type == pg.KEYDOWN:
            if event.key == pg.K_ESCAPE:
                run = False
        if event.type == pg.MOUSEBUTTONDOWN:
            mouse_pos = pg.mouse.get_pos()
            if exit_box_rect.collidepoint(mouse_pos):
                run = False
            else:
                for each in staff_list:
                    each.button_click(mouse_pos, clean_time, current_day, current_date, current_month)

pg.quit()
