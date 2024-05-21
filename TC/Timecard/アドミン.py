import os
import openpyxl as xl
from datetime import datetime


staff_list = []
current_datetime = datetime.now()
current_month = current_datetime.month
clock_in_col = 'B'
clock_out_col = 'C'
totals_col = 'D'
month_total_col = 'E'


def calculate_time_difference(workbook, sheet):
    sheet = sheet
    total_daily_mins_bin = 0
    for row in range(2, sheet.max_row + 1):
        clock_in_value = sheet[f'{clock_in_col}{row}'].value
        clock_out_value = sheet[f'{clock_out_col}{row}'].value

        if clock_in_value is not None and clock_out_value is not None:
            try:
                clean_clock_in_mins = str(clock_in_value).split(':')[1]
                clean_clock_in_hours = str(clock_in_value).split(':')[0]
                clean_clock_in_mins_as_str = int(clean_clock_in_mins)
                hours_to_mins = int(clean_clock_in_hours) * 60
                clock_in_mins = hours_to_mins + clean_clock_in_mins_as_str
                clean_clock_out_mins = str(clock_out_value).split(':')[1]
                clean_clock_out_hours = str(clock_out_value).split(':')[0]
                clean_clock_out_mins_as_str = int(clean_clock_out_mins)
                hours_to_mins = int(clean_clock_out_hours) * 60
                clock_out_mins = hours_to_mins + clean_clock_out_mins_as_str
                time_diff_in_mins = clock_out_mins - clock_in_mins
                time_diff_in_hours = time_diff_in_mins // 60
                time_diff_in_mins_remainder = time_diff_in_mins % 60
                time_string = f'{time_diff_in_hours}:{time_diff_in_mins_remainder:02d}'
                sheet[f'{totals_col}{row}'] = time_string
            except ValueError:
                sheet[f'{totals_col}{row}'] = 'Invalid time format'

    for row in range(2, sheet.max_row + 1):
        daily_total = sheet[f'{totals_col}{row}'].value
        if daily_total is not None:
            daily_total_mins_part = str(daily_total).split(':')[1]
            daily_total_hours_part = str(daily_total).split(':')[0]
            daily_total_in_mins_as_int = int(daily_total_mins_part)
            hours_to_mins = int(daily_total_hours_part) * 60
            total_daily_mins = hours_to_mins + daily_total_in_mins_as_int
            total_daily_mins_bin += total_daily_mins

    total_time_hours = total_daily_mins_bin // 60
    total_time_minutes = total_daily_mins_bin % 60
    total_cell_value = f'{total_time_hours}:{total_time_minutes:02d}'
    total_title_cell = sheet[f'{month_total_col}{2}']
    total_title_cell.value = 'Total Hours Worked'
    total_cell = sheet[f'{month_total_col}{2}']
    total_cell.value = total_cell_value

    workbook.save(filename=f'tc2assets/{current_month}.xlsx')


def perform_action_on_sheets():
    workbook = xl.load_workbook(filename=f'tc2assets/{current_month}.xlsx')
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        calculate_time_difference(workbook, sheet)
    workbook.save(filename=f'tc2assets/{current_month}.xlsx')
    print(f'合計を{current_month}.xlsxに保存しました。')


def load_staff_list():
    global staff_list
    staff_list = []
    if os.path.exists('tc2assets/staff_list.txt'):
        with open('tc2assets/staff_list.txt', 'r', encoding='utf-8') as file:
            staff_list = [line.strip() for line in file.readlines()]


def save_staff_list():
    with open('tc2assets/staff_list.txt', 'w', encoding='utf-8') as file:
        for staff in staff_list:
            file.write(f"{staff}\n")


def menu():
    print("""
==========
管理者メニュー
==========

1. スタッフの追加
2. スタッフの削除
3. スタッフの編集
4. スタッフリストを見る
5. 合計時間の計算
6. 終了する
""")
    menu_choice = input("選択を入力>　")
    if menu_choice == "1" or menu_choice == "１":
        add()
    elif menu_choice == "2" or menu_choice == "２":
        remove()
    elif menu_choice == "3" or menu_choice == "３":
        edit()
    elif menu_choice == "4" or menu_choice == "４":
        view()
    elif menu_choice == "5" or menu_choice == "５":
        perform_action_on_sheets()
    elif menu_choice == "6" or menu_choice == "６":
        global run
        run = False


def add():
    load_staff_list()
    print("""
=========
スタッフ追加
=========
ｂを入力するとメニューに戻ります。
""")
    new_name = input("名前を追加>　")
    if new_name == 'back' or new_name == 'ｂ':
        return
    if new_name in staff_list:
        print(f"{new_name}はすでにスタッフリストに存在している。")
        return
    print('確認しますか？')
    confirm = input("y/n> ")
    new_name = str(new_name)
    if confirm.lower() == 'y' or confirm == 'ｙ':
        staff_list.append(new_name + ':out')
        save_staff_list()
        print(f"{new_name.replace(':out', '')}を追加した！")
    else:
        print(f"{new_name.replace(':out', '')}は追加されなかった。")


def remove():
    load_staff_list()
    print("""
============
スタッフ削除
============
ｂを入力するとメニューに戻ります。
""")
    removed_name = input("削除する名前>　")
    if removed_name == 'back' or removed_name == 'ｂ':
        return
    if removed_name + ":out" not in staff_list:
        print(f"{removed_name}スタッフリストにはない。")
        return
    print('確認しますか？')
    confirm = input("y/n> ")
    if confirm.lower() == 'y' or confirm == 'ｙ':
        staff_list.remove(removed_name + ":out")
        save_staff_list()
        print(f"{removed_name}を取り除いた！")
    else:
        print(f"{removed_name}を削除していない。")


def edit():
    load_staff_list()
    print("""
=========
編集スタッフ
=========
ｂを入力するとメニューに戻ります。
""")
    edit_name = input("編集する名前>".replace(':out', '').replace(':in', ''))
    if edit_name == 'back' or edit_name == 'ｂ':
        return
    if edit_name not in staff_list:
        print(f"{edit_name}スタッフリストにはない。")
        return
    new_name = input(f"{edit_name}の新しい名前を入力> ")
    if new_name in staff_list:
        print(f"{new_name}はすでにスタッフリストに存在している。")
        return
    print('確認しますか？')
    confirm = input("y/n> ")
    if confirm.lower() == 'y' or confirm == 'ｙ':
        index = staff_list.index(edit_name)
        staff_list[index] = new_name + ':out'
        save_staff_list()
        print(f"{edit_name.replace(':out', '')}を{new_name.replace(':out', '')}に変更しました！")
    else:
        print(f"{edit_name.replace(':out', '')}は変更されていない。")


def view():
    load_staff_list()
    print("スタッフリスト：")
    for staff in staff_list:
        print(staff.replace(':out', ''))
    input("Enter を押してメニューに戻る。")


run = True
while run:
    menu()
